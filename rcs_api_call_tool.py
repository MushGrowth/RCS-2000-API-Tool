import json  # 导入JSON模块，用于解析和格式化报文
import queue  # 导入队列模块，用于线程间传递日志
import sys  # 导入系统模块，用于判断exe运行目录
import threading  # 导入线程模块，用于后台执行接口调用
import time  # 导入时间模块，用于暂停控制
import uuid  # 导入UUID模块，用于生成唯一编号
from datetime import datetime  # 导入日期时间模块，用于日志命名
from pathlib import Path  # 导入路径模块，用于处理文件路径
from tkinter import BOTH, END, LEFT, RIGHT, X, BooleanVar, Button, Canvas, Checkbutton, Entry, Frame, Label, Radiobutton, Scrollbar, StringVar, Tk, Toplevel, filedialog, messagebox, scrolledtext  # 导入Tkinter基础控件
from tkinter import ttk  # 导入Tkinter高级控件

import requests  # 导入HTTP请求库
from openpyxl import Workbook, load_workbook  # 导入Excel读写库

from callback_server import CallbackServer, ResponseProfile, V3_CALLBACK_PATHS, V4_CALLBACK_PATHS
from interface_catalog import load_interfaces
from rcs_http import build_v4_headers, compact_json, post_json, sign_v4_request


APP_NAME = "RCS-2000接口调用工具"  # 定义软件名称
APP_VERSION = "V1.1"  # 定义软件版本
APP_TITLE = f"{APP_NAME} {APP_VERSION}"  # 定义软件窗口标题
CONFIG_FILE_NAME = "RCS-2000接口调用工具配置.json"  # 定义配置文件名
LEGACY_CONFIG_FILE_NAME = "RCS接口调用工具配置.json"  # 定义旧版配置文件名
DEFAULT_IP = "IP"  # 定义默认IP占位符
DEFAULT_VERSION = "RCS 3.x"  # 定义默认RCS版本
VERSION_3 = "RCS 3.x"  # 定义3.x版本显示文本
VERSION_4 = "RCS 4.x"  # 定义4.x版本显示文本
TIMEOUT_SECONDS = 30  # 定义接口超时时间
NETWORK_SETTINGS = {
    "connect_timeout": 30.0,
    "read_timeout": 60.0,
    "verify_tls": True,
    "app_key": "",
    "app_secret": "",
    "source": "",
    "api_version": "v1.0",
    "algorithm": "HMAC-SHA256",
}
USED_REQ_CODES = set()  # 记录3.x本次运行已使用的reqCode
USED_4X_REQUEST_IDS = set()  # 记录4.x本次运行已使用的请求头ID
USER_MANUAL_TEXT = """RCS-2000接口调用工具 V1.0 用户使用说明

一、工具用途
本工具专为海康 RCS-2000 平台的接口调用、联调和批量数据验证设计，
支持 RCS 3.x 和 RCS 4.x 接口，支持单次调用和批量调用。
单次调用适合临时测试接口；批量调用适合通过 Excel 批量发送多行数据。

本工具适用于 RCS-2000 项目实施、接口联调、现场调试和问题排查。
实际接口地址、字段及返回结果应以当前项目所部署的 RCS-2000 版本为准。
当前软件版本：V1.0。

二、版本选择
顶部 RCS版本 可选择：
1. RCS 3.x
2. RCS 4.x

切换版本后：
1. 当前字段配置会清空。
2. 接口下拉框会切换为对应版本的预置接口。
3. 程序会自动选择该版本第一个接口并加载字段。
4. 端口和完整URL会按版本规则刷新。

三、URL规则
RCS 3.x:
http://IP:8182/rcms/services/rest/hikRpcService/接口方法名

RCS 4.x:
https://IP:443/rcs/rtas/接口路径后缀

IP、端口、接口后缀和完整URL都可以手动修改，最终调用以完整URL为准。

四、调用方式
调用方式下拉框有两个选项：
1. 单次
2. 批量

单次模式：
在字段配置右侧“字段值”中填写参数值，点击开始执行后只调用一次接口。

批量模式：
选择 Excel 文件，程序按字段配置中的 Excel列号逐行读取数据，每行调用一次接口。

五、选择接口
选择接口下拉框会根据当前 RCS 版本展示预置接口。
选择接口后会自动填充：
1. 接口后缀或方法名
2. 完整URL
3. 字段配置
4. 字段必填提示
5. 嵌套字段层级

自动填充后仍可手动调整字段名、列号、嵌套层级、字段顺序和字段值。

六、字段配置
字段配置每行包含：
1. 拖动图标：按住“☰”后上下拖动字段顺序。
2. Excel列号：批量模式下从 Excel 哪一列取值。
3. 有嵌套：字段是否属于对象或数组对象。
4. 嵌套层级：例如 targetRoute 或 data[0]。
5. 字段名：普通字段名或嵌套字段的子字段名。
6. 要求：只读，展示必填、条件必填或选填。
7. 字段值：单次模式下直接填写参数值。

修改 Excel列号 后，字段行会自动移动到对应位置，其他字段的列号依次顺延。

七、必填规则
必填：为空会拦截执行。
条件必填：只做提示，不强制拦截。
选填：为空不会拦截，也不会发送到请求报文中。

八、嵌套字段
对象字段示例：
targetRoute.type
界面配置：
勾选有嵌套，嵌套层级填 targetRoute，字段名填 type。

数组对象字段示例：
data[0].robotTaskCode
界面配置：
勾选有嵌套，嵌套层级填 data[0]，字段名填 robotTaskCode。

[0] 表示数组第一个对象；如需第二个对象可改为 [1]。

九、JSON数组或对象
普通字段也可以直接填写 JSON 字符串。
例如 robots 字段可填：
["R001","R002"]

extra 字段可填：
{"key":"value"}

程序会自动识别 JSON 数组或对象。

十、执行控制
开始执行：开始单次或批量调用。
暂停执行：批量调用中暂停后续请求，当前请求会等待返回。
继续执行：暂停后开始按钮会变为继续执行。
停止执行：停止后续请求，下次开始会重新从第一行执行。

十一、成功判断
以下任一情况视为成功：
1. code == 0
2. code == "0"
3. code == "SUCCESS"

十二、3.x和4.x差异
RCS 3.x：
请求报文会自动加入 reqCode。

RCS 4.x：
请求报文不发送 reqCode。
请求头会自动加入 X-lr-request-id，最大16字节，每次调用唯一。

十三、日志和失败行
界面会显示发送报文、应答报文和执行结果。
日志默认只显示在界面中，不会自动保存到本地。
如需保存，可点击“导出日志”选择路径并导出为 TXT 文件。
失败行会进入失败列表，点击后可查看该行发送报文和应答报文。
可点击“清空失败行”清除当前失败列表和失败明细。
可点击导出失败行，将失败数据导出为 Excel。

十四、模板导出
配置字段后点击导出Excel模板，会生成带字段名表头的 Excel。
模板列顺序与当前字段顺序一致。

十五、配置保存
关闭软件时会自动保存：
RCS版本、IP、端口、接口后缀、完整URL、选择接口、调用方式、Excel路径、字段配置、字段值等。
下次打开同一 exe 会自动恢复。

十六、隐私与环境信息
软件说明采用通用描述，不包含开发者的用户名、盘符或本机绝对路径。
运行时填写的 IP、Excel路径和字段值仅用于当前工具配置及接口调用。
"""  # 定义内置用户使用说明文本


PRESET_APIS_3X = [  # 定义RCS3.x预置接口
    {"display_name": "生成任务单", "suffix": "genAgvSchedulingTask", "fields": [("taskTyp", "必填"), ("wbCode", "条件必填-与positionCodePath二选一"), ("positionCodePath[0].positionCode", "条件必填"), ("positionCodePath[0].type", "条件必填"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填"), ("podCode", "选填"), ("podDir", "选填"), ("podTyp", "选填"), ("materialLot", "选填"), ("materialType", "选填-移载机器人专用"), ("priority", "选填"), ("taskCode", "选填"), ("agvCode", "选填"), ("groupId", "选填"), ("agvTyp", "选填"), ("positionSelStrategy", "选填"), ("data", "选填")]},  # 定义生成任务单
    {"display_name": "继续执行任务", "suffix": "continueTask", "fields": [("wbCode", "条件必填-四选一"), ("podCode", "条件必填-四选一"), ("agvCode", "条件必填-四选一"), ("taskCode", "条件必填-四选一"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填"), ("taskSeq", "选填"), ("nextPositionCode.positionCode", "选填-外部设置时需要"), ("nextPositionCode.type", "选填")]},  # 定义继续执行任务
    {"display_name": "取消任务", "suffix": "cancelTask", "fields": [("agvCode", "条件必填-与taskCode二选一"), ("taskCode", "条件必填-与agvCode二选一"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填"), ("forceCancel", "选填-默认为0"), ("matterArea", "选填")]},  # 定义取消任务
    {"display_name": "任务优先级设置", "suffix": "setTaskPriority", "fields": [("priorities[0].taskCode", "必填"), ("priorities[0].priority", "必填"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填")]},  # 定义任务优先级设置
    {"display_name": "货架与位置绑定、解绑", "suffix": "bindPodAndBerth", "fields": [("podCode", "必填"), ("positionCode", "必填"), ("indBind", "必填-1绑定/0解绑"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填"), ("podDir", "选填"), ("characterValue", "选填")]},  # 定义货架与位置绑定解绑
    {"display_name": "货架与物料绑定、解绑", "suffix": "bindPodAndMat", "fields": [("podCode", "必填"), ("materialLot", "必填"), ("indBind", "必填-1绑定/0解绑"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填")]},  # 定义货架与物料绑定解绑
    {"display_name": "位置禁用与启用", "suffix": "lockPosition", "fields": [("positionCode", "必填"), ("indBind", "必填-1启用/0禁用"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填")]},  # 定义位置禁用启用
    {"display_name": "地图位置信息同步", "suffix": "syncMapDatas", "fields": [("mapShortName", "必填"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填"), ("mapDataCode", "选填"), ("dataTyp", "选填")]},  # 定义地图同步
    {"display_name": "查询货架储位与物料批次关系", "suffix": "queryPodBerthAndMat", "fields": [("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填"), ("podCode", "选填"), ("materialLot", "选填"), ("positionCode", "选填"), ("areaCode", "选填"), ("mapShortName", "选填")]},  # 定义查询货架储位物料
    {"display_name": "仓位禁用与启用", "suffix": "blockStgBin", "fields": [("data[0].stgBinCode", "必填"), ("data[0].action", "必填-0启用/1禁用"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填")]},  # 定义仓位禁用启用
    {"display_name": "容器与仓位绑定、解绑", "suffix": "bindCtnrAndBin", "fields": [("ctnrTyp", "必填"), ("indBind", "必填-1绑定/0解绑"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填"), ("ctnrCode", "选填"), ("stgBinCode", "选填"), ("binName", "选填"), ("characterValue", "选填"), ("positionCode", "选填")]},  # 定义容器仓位绑定解绑
    {"display_name": "查询任务状态", "suffix": "queryTaskStatus", "fields": [("taskCodes", "条件必填-与agvCode二选一"), ("agvCode", "条件必填-与taskCodes二选一"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填")]},  # 定义查询任务状态
    {"display_name": "查询 AGV 状态", "suffix": "queryAgvStatus", "fields": [("mapCode", "必填"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填")]},  # 定义查询AGV状态
    {"display_name": "停止 AGV", "suffix": "stopRobot", "fields": [("mapShortName", "条件必填-robotCount=-1时必填"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填"), ("robotCount", "选填-填-1时代表全部"), ("robots", "选填-具体车号列表")]},  # 定义停止AGV
    {"display_name": "恢复 AGV", "suffix": "resumeRobot", "fields": [("mapShortName", "条件必填-robotCount=-1时必填"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填"), ("robotCount", "选填-填-1时代表全部"), ("robots", "选填-具体车号列表")]},  # 定义恢复AGV
    {"display_name": "区域清空/释放", "suffix": "blockArea", "fields": [("matterArea", "必填"), ("indBind", "必填-1封锁/0解封"), ("targetArea", "条件必填-controlMod=2时必填"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填"), ("pause", "选填-0不暂停/1暂停"), ("controlMod", "选填"), ("noticeThird", "选填")]},  # 定义区域清空释放
    {"display_name": "预调度对外接口", "suffix": "genPreScheduleTask", "fields": [("positionCode", "必填"), ("nextTask", "必填"), ("agvTyp", "必填"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填"), ("priority", "选填"), ("useableLayers", "选填-CTU用"), ("cacheCount", "选填-CTU用"), ("update", "选填-0不更新/1更新")]},  # 定义预调度接口
    {"display_name": "清空巷道", "suffix": "clearRoadWay", "fields": [("reqTime", "必填-此接口特殊"), ("clientCode", "必填"), ("tokenCode", "必填"), ("roadWayCode", "必填")]},  # 定义清空巷道
    {"display_name": "料箱出库TPS（CTU+分拨墙）", "suffix": "getOutPod", "fields": [("reqTime", "必填"), ("clientCode", "必填"), ("tokenCode", "必填"), ("taskTyp", "必填-传1"), ("data[0].taskCode", "必填"), ("data[0].ctnrCode", "必填"), ("data[0].wbCode", "必填"), ("data[0].binCode", "选填"), ("data[0].agvTyp", "选填"), ("data[0].priority", "选填")]},  # 定义料箱出库TPS
    {"display_name": "料箱回库TPS（CTU+分拨墙）", "suffix": "returnPod", "fields": [("reqTime", "必填"), ("clientCode", "必填"), ("tokenCode", "必填"), ("taskCode", "必填"), ("taskTyp", "必填-传5"), ("ctnrCode", "必填"), ("binCode", "必填"), ("srcBinCode", "必填"), ("wbCode", "条件必填-初始化入库时必填"), ("agvTyp", "选填"), ("returnPodStrategy", "选填")]},  # 定义料箱回库TPS
    {"display_name": "料箱顺序出库（CTU）", "suffix": "genCtuGroupTaskBatch", "fields": [("taskTyp", "必填-默认B10"), ("seqTyp", "必填"), ("taskGroups[0].taskCode", "必填"), ("taskGroups[0].ctnrCode", "必填"), ("taskGroups[0].ctnrTyp", "必填"), ("taskGroups[0].wbCode", "必填"), ("taskGroups[0].wbTyp", "必填"), ("taskGroups[0].groupId", "必填"), ("taskGroups[0].sequence", "必填"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填")]},  # 定义料箱顺序出库
    {"display_name": "料箱取放回调（CTU）", "suffix": "boxApplyPass", "fields": [("reqTime", "必填"), ("clientCode", "必填"), ("tokenCode", "必填"), ("taskCode", "必填"), ("type", "必填-1取通过/2放通过")]},  # 定义料箱取放回调
    {"display_name": "物料类型与位置绑定、解绑（移载机器人协议）", "suffix": "bindMaterialTypCodeAndBerth", "fields": [("materialType", "条件必填-绑定时必填,解绑可不填"), ("positionCode", "必填"), ("indBind", "必填-1绑定/0解绑"), ("reqTime", "选填"), ("clientCode", "选填"), ("tokenCode", "选填"), ("ctnrCode", "选填")]},  # 定义物料类型位置绑定解绑
]  # 结束RCS3.x预置接口


PRESET_APIS_4X = [  # 定义RCS4.x预置接口
    {"display_name": "任务组接口", "suffix": "api/robot/controller/task/group", "fields": [("groupCode", "必填"), ("strategy", "必填"), ("targetRoute.type", "必填"), ("targetRoute.code", "必填"), ("data[0].robotTaskCode", "必填"), ("data[0].sequence", "必填"), ("strategyValue", "条件必填"), ("groupSeq", "选填")]},  # 定义任务组接口
    {"display_name": "任务下发接口", "suffix": "api/robot/controller/task/submit", "fields": [("taskType", "必填"), ("targetRoute[0].type", "必填"), ("targetRoute[0].code", "必填"), ("targetRoute[0].operation", "选填"), ("targetRoute[0].robotType", "选填"), ("targetRoute[0].robotCode", "选填"), ("targetRoute[0].extra", "选填"), ("initPriority", "选填"), ("deadline", "选填"), ("expectedStartTime", "选填"), ("robotType", "选填"), ("robotCode", "选填"), ("interrupt", "选填"), ("robotTaskCode", "选填"), ("groupCode", "选填"), ("extra", "选填")]},  # 定义任务下发接口
    {"display_name": "任务继续执行接口", "suffix": "api/robot/controller/task/extend/continue", "fields": [("triggerType", "必填"), ("triggerCode", "必填"), ("targetRoute.type", "选填"), ("targetRoute.code", "选填"), ("targetRoute.operation", "选填"), ("targetRoute.robotType", "选填"), ("targetRoute.robotCode", "选填"), ("targetRoute.extra", "选填"), ("extra", "选填")]},  # 定义任务继续执行接口
    {"display_name": "任务取消接口", "suffix": "api/robot/controller/task/cancel", "fields": [("cancelType", "必填"), ("returnTaskType", "条件必填"), ("robotTaskCode", "选填"), ("carrierCode", "选填"), ("robotCode", "选填"), ("reason", "选填"), ("targetRoute.type", "选填"), ("targetRoute.code", "选填"), ("extra", "选填")]},  # 定义任务取消接口
    {"display_name": "任务优先级设置接口", "suffix": "api/robot/controller/task/priority", "fields": [("robotTaskCode", "必填"), ("initPriority", "必填"), ("deadline", "选填"), ("extra", "选填")]},  # 定义任务优先级设置接口
    {"display_name": "区域暂停与恢复机器人接口", "suffix": "api/robot/controller/zone/pause", "fields": [("zoneCode", "必填"), ("invoke", "必填"), ("mapCode", "选填"), ("extra", "选填")]},  # 定义区域暂停恢复接口
    {"display_name": "区域归巢机器人接口", "suffix": "api/robot/controller/zone/homing", "fields": [("autoShutdown", "必填"), ("bootTime", "条件必填"), ("mapCode", "选填"), ("zoneCode", "选填"), ("zoneCodes", "选填"), ("expireTime", "选填"), ("extra", "选填")]},  # 定义区域归巢接口
    {"display_name": "区域驱离机器人接口", "suffix": "api/robot/controller/zone/banish", "fields": [("zoneCode", "选填"), ("zoneCodes", "选填"), ("targetZoneCode", "选填"), ("pause", "选填"), ("report", "选填"), ("controlMode", "选填"), ("expireTime", "选填"), ("invoke", "选填"), ("extra", "选填")]},  # 定义区域驱离接口
    {"display_name": "区域封锁与恢复接口", "suffix": "api/robot/controller/zone/blockade", "fields": [("zoneCode", "必填"), ("invoke", "必填"), ("mapCode", "选填"), ("pause", "选填"), ("report", "选填"), ("disableSite", "选填"), ("enableSite", "选填"), ("extra", "选填")]},  # 定义区域封锁恢复接口
    {"display_name": "载具与站点绑定接口", "suffix": "api/robot/controller/carrier/bind", "fields": [("carrierCode", "必填"), ("siteCode", "必填"), ("carrierDir", "选填"), ("extra", "选填")]},  # 定义载具站点绑定接口
    {"display_name": "载具与站点解绑接口", "suffix": "api/robot/controller/carrier/unbind", "fields": [("carrierCode", "选填"), ("siteCode", "选填"), ("extra", "选填")]},  # 定义载具站点解绑接口
    {"display_name": "存储对象与搬运对象绑定解绑接口", "suffix": "api/robot/controller/site/bind", "fields": [("invoke", "必填"), ("slotCategory", "条件必填"), ("slotCode", "条件必填"), ("carrierCategory", "选填"), ("carrierType", "选填"), ("carrierCode", "选填"), ("carrierDir", "选填"), ("stackLabel", "选填"), ("colCount", "选填"), ("extra", "选填")]},  # 定义存储对象绑定解绑接口
    {"display_name": "载具禁用与启用", "suffix": "api/robot/controller/carrier/lock", "fields": [("carrierCode", "必填"), ("invoke", "必填")]},  # 定义载具禁用启用接口
    {"display_name": "站点禁用与启用", "suffix": "api/robot/controller/site/lock", "fields": [("siteCode", "必填"), ("invoke", "必填")]},  # 定义站点禁用启用接口
    {"display_name": "外设执行通知接口【返回0】", "suffix": "spi/wcs/robot/eqpt/notify", "fields": [("eqptCode", "必填"), ("taskCode", "必填"), ("actionStatus", "必填"), ("siteCode", "选填"), ("carrierInfo", "选填"), ("extra", "选填")]},  # 定义外设通知返回0接口
    {"display_name": "预调度任务下发接口", "suffix": "api/robot/controller/task/pretask", "fields": [("siteCode", "必填"), ("nextTaskTime", "必填"), ("robotType", "选填"), ("priority", "选填"), ("taskCount", "选填"), ("capacityCount", "选填"), ("amrDir", "选填"), ("extra", "选填")]},  # 定义预调度任务接口
    {"display_name": "查询任务状态接口", "suffix": "api/robot/controller/task/query", "fields": [("robotTaskCode", "必填")]},  # 定义查询任务状态接口
    {"display_name": "查询机器人状态接口", "suffix": "api/robot/controller/robot/query", "fields": [("singleRobotCode", "必填")]},  # 定义查询机器人状态接口
    {"display_name": "查询载具状态接口", "suffix": "api/robot/controller/carrier/query", "fields": [("carrierCode", "必填")]},  # 定义查询载具状态接口
    {"display_name": "物料绑定接口", "suffix": "api/robot/controller/matlabel/bind", "fields": [("carrierCode", "必填"), ("matLabel", "必填")]},  # 定义物料绑定接口
    {"display_name": "物料解绑接口", "suffix": "api/robot/controller/matlabel/unbind", "fields": [("carrierCode", "必填"), ("matLabel", "必填")]},  # 定义物料解绑接口
    {"display_name": "外设执行通知接口【返回SUCCESS】", "suffix": "spi/wcs/robot/eqpt/notifyGbt", "fields": [("eqptCode", "必填"), ("taskCode", "必填"), ("actionStatus", "必填"), ("siteCode", "选填"), ("carrierInfo", "选填"), ("extra", "选填")]},  # 定义外设通知返回SUCCESS接口
]  # 结束RCS4.x预置接口


def normalize_cell(value) -> str:  # 定义单元格内容标准化函数
    if value is None:  # 判断值是否为空
        return ""  # 空值返回空字符串
    return str(value).strip()  # 转为字符串并去除空白


def column_index_to_name(index: int) -> str:  # 定义列号转Excel字母函数
    name = ""  # 初始化列名
    while index > 0:  # 循环计算列名
        index, remainder = divmod(index - 1, 26)  # 计算当前字符
        name = chr(65 + remainder) + name  # 拼接列名字母
    return name  # 返回列名


def requirement_sort_key(requirement: str) -> int:  # 定义必填排序规则
    if requirement.startswith("必填"):  # 判断是否必填
        return 0  # 必填排第一
    if requirement.startswith("条件必填"):  # 判断是否条件必填
        return 1  # 条件必填排第二
    return 2  # 选填排最后


def is_strict_required(requirement: str) -> bool:  # 定义严格必填判断
    return requirement.startswith("必填")  # 只有必填才强制校验


def parse_cell_payload_value(value: str):  # 定义字段值解析函数
    text = normalize_cell(value)  # 标准化字段值
    if not text:  # 判断是否为空
        return ""  # 空值直接返回
    if text[0] in "[{":  # 判断是否可能是JSON
        try:  # 尝试解析JSON
            return json.loads(text)  # 返回JSON对象
        except json.JSONDecodeError:  # 捕获JSON解析失败
            return text  # 失败时按字符串返回
    return text  # 普通值按字符串返回


def parse_field_path(field_name: str) -> list:  # 定义字段路径解析函数
    tokens = []  # 初始化路径片段
    current = ""  # 初始化当前片段
    index = 0  # 初始化字符位置
    while index < len(field_name):  # 循环解析字段路径
        char = field_name[index]  # 获取当前字符
        if char == ".":  # 判断对象分隔符
            if current:  # 判断当前片段是否存在
                tokens.append(current)  # 保存当前片段
                current = ""  # 清空当前片段
            index += 1  # 移动位置
            continue  # 继续解析
        if char == "[":  # 判断数组下标开始
            if current:  # 判断当前片段是否存在
                tokens.append(current)  # 保存当前片段
                current = ""  # 清空当前片段
            end_index = field_name.find("]", index)  # 查找数组下标结束
            if end_index == -1:  # 判断是否缺少结束符
                return tokens  # 返回已解析片段
            index_text = field_name[index + 1:end_index].strip()  # 获取下标文本
            tokens.append(int(index_text) if index_text.isdigit() else index_text)  # 保存数组下标
            index = end_index + 1  # 移动到结束符后
            continue  # 继续解析
        current += char  # 累加普通字段字符
        index += 1  # 移动位置
    if current:  # 判断最后片段是否存在
        tokens.append(current)  # 保存最后片段
    return tokens  # 返回路径片段


def ensure_list_size(target: list, index: int) -> None:  # 定义确保列表长度函数
    while len(target) <= index:  # 判断长度是否不足
        target.append({})  # 追加空对象


def set_nested_value(target: dict, field_name: str, value) -> None:  # 定义设置嵌套字段值函数
    tokens = parse_field_path(field_name)  # 解析字段路径
    current = target  # 设置当前节点
    for position, token in enumerate(tokens):  # 遍历路径片段
        is_last = position == len(tokens) - 1  # 判断是否最后片段
        next_token = None if is_last else tokens[position + 1]  # 获取下个片段
        if isinstance(token, int):  # 判断是否数组下标
            if not isinstance(current, list):  # 判断当前是否不是列表
                return  # 类型不匹配则跳过
            ensure_list_size(current, token)  # 确保列表长度
            if is_last:  # 判断是否最后片段
                current[token] = value  # 设置数组值
            else:  # 处理中间片段
                if not isinstance(current[token], (dict, list)):  # 判断子节点类型
                    current[token] = [] if isinstance(next_token, int) else {}  # 创建容器
                current = current[token]  # 进入子节点
            continue  # 继续循环
        if is_last:  # 判断是否最后片段
            current[token] = value  # 设置对象字段值
            continue  # 继续循环
        if token not in current or not isinstance(current[token], (dict, list)):  # 判断子节点是否存在
            current[token] = [] if isinstance(next_token, int) else {}  # 创建子容器
        current = current[token]  # 进入子节点


def build_payload_values(flat_values: dict) -> dict:  # 定义构造请求业务字段函数
    payload_values = {}  # 初始化请求字段对象
    for field_name, raw_value in flat_values.items():  # 遍历扁平字段
        if normalize_cell(raw_value) == "":  # 判断是否空值
            continue  # 空值不写入报文
        parsed_value = parse_cell_payload_value(raw_value)  # 解析字段值
        if "." in field_name or "[" in field_name:  # 判断是否嵌套字段
            set_nested_value(payload_values, field_name, parsed_value)  # 写入嵌套字段
        else:  # 处理普通字段
            payload_values[field_name] = parsed_value  # 写入普通字段
    return payload_values  # 返回请求字段对象


def split_field_for_ui(field_name: str) -> tuple[bool, str, str]:  # 定义字段路径拆分函数
    text = field_name.strip()  # 清理字段名
    if "." not in text:  # 判断是否普通字段
        return False, "", text  # 普通字段不嵌套
    nested_path, child_name = text.rsplit(".", 1)  # 拆分父层级和字段名
    return True, nested_path, child_name  # 返回嵌套字段配置


def combine_nested_field_name(nested_path: str, field_name: str) -> str:  # 定义组合嵌套字段名函数
    parent_path = nested_path.strip().strip(".")  # 清理父层级
    child_name = field_name.strip().strip(".")  # 清理字段名
    return f"{parent_path}.{child_name}" if parent_path else child_name  # 返回完整字段名


def get_config_path() -> Path:  # 定义获取配置文件路径函数
    if getattr(sys, "frozen", False):  # 判断是否exe运行
        return Path(sys.executable).with_name(CONFIG_FILE_NAME)  # exe同目录配置
    return Path(__file__).with_name(CONFIG_FILE_NAME)  # 源码同目录配置


def get_legacy_config_path() -> Path:  # 定义获取旧版配置文件路径函数
    if getattr(sys, "frozen", False):  # 判断是否exe运行
        return Path(sys.executable).with_name(LEGACY_CONFIG_FILE_NAME)  # 返回exe同目录旧配置
    return Path(__file__).with_name(LEGACY_CONFIG_FILE_NAME)  # 返回源码同目录旧配置


def get_resource_path(relative_path: str) -> Path:  # 定义获取程序资源路径函数
    base_path = Path(getattr(sys, "_MEIPASS", Path(__file__).parent))  # 获取打包资源目录或源码目录
    return base_path / relative_path  # 返回资源完整路径


def build_req_code() -> str:  # 定义3.x reqCode生成函数
    while True:  # 循环直到不重复
        req_code = uuid.uuid4().hex  # 生成32位UUID
        if req_code not in USED_REQ_CODES:  # 判断是否未使用
            USED_REQ_CODES.add(req_code)  # 记录已使用
            return req_code  # 返回reqCode


def build_4x_request_id() -> str:  # 定义4.x请求头ID生成函数
    while True:  # 循环直到不重复
        request_id = uuid.uuid4().hex[:16]  # 生成16字节以内ID
        if request_id not in USED_4X_REQUEST_IDS:  # 判断是否未使用
            USED_4X_REQUEST_IDS.add(request_id)  # 记录已使用
            return request_id  # 返回请求头ID


def get_presets(version: str) -> list[dict]:  # 定义按版本获取预置接口函数
    return load_interfaces(version)  # 从版本化JSON配置加载接口


def get_base_url(version: str, ip: str, port: str) -> str:  # 定义生成基础URL函数
    if version == VERSION_4:  # 判断是否4.x
        return f"https://{ip}:{port}/rcs/rtas/"  # 返回4.x基础URL
    return f"http://{ip}:{port}/rcms/services/rest/hikRpcService/"  # 返回3.x基础URL


def is_success_response(response_json: dict) -> bool:  # 定义成功响应判断函数
    code = response_json.get("code")  # 获取响应code
    return code == 0 or code == "0" or code == "SUCCESS"  # 兼容3.x和4.x成功码


def read_excel_rows(excel_path: Path, field_configs: list[dict]) -> list[dict]:  # 定义读取Excel函数
    workbook = load_workbook(excel_path, read_only=True, data_only=True)  # 只读打开Excel
    worksheet = workbook.active  # 获取活动表
    rows = []  # 初始化数据行
    try:  # 确保关闭工作簿
        for row_index, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True), start=1):  # 遍历Excel行
            values = {}  # 初始化当前行字段值
            for config in field_configs:  # 遍历字段配置
                column_index = config["column_index"] - 1  # 转为Python下标
                field_name = config["field_name"]  # 获取字段名
                values[field_name] = normalize_cell(row[column_index] if len(row) > column_index else None)  # 读取单元格
            if not any(values.values()):  # 判断是否空行
                continue  # 跳过空行
            if row_index == 1 and any(values.get(config["field_name"]) == config["field_name"] for config in field_configs):  # 判断表头
                continue  # 跳过表头
            rows.append({"row_index": row_index, "values": values})  # 保存数据行
    finally:  # 最终清理
        workbook.close()  # 关闭工作簿
    return rows  # 返回数据行


def validate_row(row: dict, field_configs: list[dict]) -> str:  # 定义行数据校验函数
    for config in field_configs:  # 遍历字段配置
        field_name = config["field_name"]  # 获取字段名
        requirement = config.get("requirement", "选填")  # 获取字段要求
        if is_strict_required(requirement) and not row["values"].get(field_name):  # 判断必填为空
            return f"字段{field_name}为必填，当前为空"  # 返回错误
    values = row["values"]
    if values.get("method") == "APPLY_RESOURCE":
        for field_name in ("srcCode", "syncMode"):
            if not values.get(field_name):
                return f"字段{field_name}在交管区申请时为必填"
    return ""  # 返回无错误


def post_api_request(version: str, api_url: str, row: dict) -> tuple[bool, str, dict, str]:  # 定义接口调用函数
    payload = build_payload_values(row["values"])  # 构造业务报文
    headers = {"Content-Type": "application/json;charset=UTF-8"}  # 初始化请求头
    signing_source = ""
    if version == VERSION_3:  # 判断3.x
        payload = {"reqCode": build_req_code(), **payload}  # 3.x自动加入reqCode
    else:  # 处理4.x
        headers = build_v4_headers(
            build_4x_request_id(),
            NETWORK_SETTINGS["app_key"],
            NETWORK_SETTINGS["source"],
            NETWORK_SETTINGS["api_version"],
            algorithm=NETWORK_SETTINGS["algorithm"],
        )
        if NETWORK_SETTINGS["app_key"] and NETWORK_SETTINGS["app_secret"]:
            api_url, signing_source = sign_v4_request(
                "POST", api_url, headers, compact_json(payload), NETWORK_SETTINGS["app_secret"]
            )
    response_text = ""  # 初始化响应文本
    try:  # 捕获请求异常
        response = post_json(
            api_url,
            payload,
            headers,
            NETWORK_SETTINGS["connect_timeout"],
            NETWORK_SETTINGS["read_timeout"],
            NETWORK_SETTINGS["verify_tls"],
        )
        response_body = response.text.strip()  # 获取响应正文
        response_text = response_body
        response.raise_for_status()  # 检查HTTP状态
        response_json = response.json()  # 解析JSON响应
        response_text = (
            f"HTTP {response.status_code} {response.reason}\n"
            + "\n".join(f"{key}: {value}" for key, value in response.headers.items())
            + f"\n\n{response_body}"
        )
    except requests.RequestException as error:  # 捕获请求异常
        return False, f"{error.__class__.__name__}: {error}", payload, response_text  # 返回失败
    except json.JSONDecodeError as error:  # 捕获JSON解析异常
        return False, f"响应不是合法JSON：{error}，原始响应：{response_text}", payload, response_text  # 返回失败
    if is_success_response(response_json):  # 判断是否成功
        return True, "成功", payload, response_text  # 返回成功
    return False, f"接口返回code={response_json.get('code')}", payload, response_text  # 返回失败


class RcsApiCallTool:  # 定义主界面类
    def __init__(self, root: Tk) -> None:  # 定义初始化函数
        self.root = root  # 保存主窗口
        self.version_var = StringVar(value=DEFAULT_VERSION)  # 创建版本变量
        self.ip_var = StringVar(value=DEFAULT_IP)  # 创建IP变量
        self.port_var = StringVar(value="8182")  # 创建端口变量
        self.suffix_var = StringVar(value="")  # 创建接口后缀变量
        self.api_url_var = StringVar(value="")  # 创建完整URL变量
        self.selected_api_var = StringVar(value="")  # 创建接口选择变量
        self.call_mode_var = StringVar(value="单次")  # 创建调用方式变量
        self.excel_path_var = StringVar(value="")  # 创建Excel路径变量
        self.status_var = StringVar(value="请选择接口并配置字段。")  # 创建状态变量
        self.log_queue = queue.Queue()  # 创建日志队列
        self.field_rows = []  # 创建字段行列表
        self.failed_records = {}  # 创建失败记录字典
        self.is_running = False  # 初始化运行状态
        self.is_paused = False  # 初始化暂停状态
        self.pause_event = threading.Event()  # 创建暂停事件
        self.stop_event = threading.Event()  # 创建停止事件
        self.callback_server = None  # 本地回调模拟服务
        self.pause_event.set()  # 默认允许执行
        self.configure_styles()  # 配置界面样式
        self.build_ui()  # 构建界面
        self.load_config_or_default()  # 加载配置或默认接口
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)  # 绑定关闭事件
        self.root.after(100, self.flush_log_queue)  # 定时刷新日志

    def configure_styles(self) -> None:  # 定义界面样式配置函数
        self.root.configure(bg="#f3f6fb")  # 设置窗口背景色
        self.root.option_add("*Font", ("Microsoft YaHei UI", 9))  # 设置全局字体
        self.root.option_add("*Background", "#f3f6fb")  # 设置控件背景色
        self.root.option_add("*Foreground", "#24324a")  # 设置控件文字色
        self.root.option_add("*Entry.Background", "#ffffff")  # 设置输入框背景色
        self.root.option_add("*Text.Background", "#ffffff")  # 设置文本框背景色
        self.root.option_add("*Button.Background", "#e7eef9")  # 设置按钮背景色
        self.root.option_add("*Button.ActiveBackground", "#d7e5f8")  # 设置按钮按下背景色
        self.root.option_add("*Button.Relief", "flat")  # 设置按钮扁平样式
        self.root.option_add("*Button.Cursor", "hand2")  # 设置按钮鼠标样式
        self.root.option_add("*Button.Padx", 8)  # 设置按钮横向留白
        self.root.option_add("*Button.Pady", 4)  # 设置按钮纵向留白
        style = ttk.Style(self.root)  # 创建ttk样式
        style.theme_use("clam")  # 使用可定制主题
        style.configure("TCombobox", fieldbackground="#ffffff", background="#e7eef9", padding=4)  # 设置下拉框样式
        style.configure("Treeview", background="#ffffff", fieldbackground="#ffffff", foreground="#24324a", rowheight=27, borderwidth=0)  # 设置表格样式
        style.configure("Treeview.Heading", background="#dfe9f7", foreground="#1d3557", font=("Microsoft YaHei UI", 9, "bold"), padding=6)  # 设置表头样式
        style.map("Treeview", background=[("selected", "#cfe2ff")], foreground=[("selected", "#17365d")])  # 设置表格选中样式

    def build_ui(self) -> None:  # 定义构建界面函数
        self.root.title(APP_TITLE)  # 设置窗口标题
        icon_path = get_resource_path("assets/rcs2000-icon.ico")  # 获取程序图标路径
        if icon_path.exists():  # 判断图标是否存在
            self.root.iconbitmap(default=str(icon_path))  # 设置窗口和任务栏图标
        self.root.geometry("1360x900")  # 设置窗口大小
        self.root.minsize(1180, 760)  # 设置窗口最小尺寸
        top_frame = Frame(self.root)  # 创建顶部区域
        top_frame.pack(fill=X, padx=10, pady=6)  # 布局顶部区域
        Button(top_frame, text="使用说明", command=self.show_user_manual, width=12).pack(side=RIGHT)  # 创建右上角使用说明按钮
        Button(top_frame, text="回调模拟", command=self.show_callback_server, width=12).pack(side=RIGHT, padx=6)
        Button(top_frame, text="网络/签名", command=self.show_network_settings, width=12).pack(side=RIGHT)
        Label(top_frame, text="RCS版本：").pack(side=LEFT)  # 创建版本标签
        self.version_combo = ttk.Combobox(top_frame, textvariable=self.version_var, values=[VERSION_3, VERSION_4], state="readonly", width=12)  # 创建版本下拉框
        self.version_combo.pack(side=LEFT, padx=(4, 12))  # 布局版本下拉框
        self.version_combo.bind("<<ComboboxSelected>>", self.on_version_selected)  # 绑定版本切换事件
        Label(top_frame, text="调用方式：").pack(side=LEFT)  # 创建调用方式标签
        self.call_mode_combo = ttk.Combobox(top_frame, textvariable=self.call_mode_var, values=["单次", "批量"], state="readonly", width=8)  # 创建调用方式下拉框
        self.call_mode_combo.pack(side=LEFT, padx=(4, 12))  # 布局调用方式下拉框
        self.call_mode_combo.bind("<<ComboboxSelected>>", lambda _event: self.update_call_mode_ui())  # 绑定调用方式变化事件
        Label(top_frame, text="选择接口：").pack(side=LEFT)  # 创建接口标签
        self.api_combo = ttk.Combobox(top_frame, textvariable=self.selected_api_var, state="readonly", width=44)  # 创建接口下拉框
        self.api_combo.pack(side=LEFT, padx=(4, 12))  # 布局接口下拉框
        self.api_combo.bind("<<ComboboxSelected>>", self.on_preset_api_selected)  # 绑定接口选择事件
        url_frame = Frame(self.root)  # 创建URL区域
        url_frame.pack(fill=X, padx=10, pady=4)  # 布局URL区域
        Label(url_frame, text="IP：").pack(side=LEFT)  # 创建IP标签
        Entry(url_frame, textvariable=self.ip_var, width=18).pack(side=LEFT, padx=(4, 10))  # 创建IP输入框
        Label(url_frame, text="端口：").pack(side=LEFT)  # 创建端口标签
        Entry(url_frame, textvariable=self.port_var, width=8).pack(side=LEFT, padx=(4, 10))  # 创建端口输入框
        Label(url_frame, text="接口后缀/方法名：").pack(side=LEFT)  # 创建后缀标签
        Entry(url_frame, textvariable=self.suffix_var, width=42).pack(side=LEFT, padx=(4, 10))  # 创建后缀输入框
        url_full_frame = Frame(self.root)  # 创建完整URL区域
        url_full_frame.pack(fill=X, padx=10, pady=4)  # 布局完整URL区域
        Label(url_full_frame, text="完整URL：").pack(side=LEFT)  # 创建完整URL标签
        Entry(url_full_frame, textvariable=self.api_url_var).pack(side=LEFT, fill=X, expand=True, padx=(4, 0))  # 创建完整URL输入框
        self.ip_var.trace_add("write", self.update_api_url_from_parts)  # 监听IP变化
        self.port_var.trace_add("write", self.update_api_url_from_parts)  # 监听端口变化
        self.suffix_var.trace_add("write", self.update_api_url_from_parts)  # 监听后缀变化
        file_frame = Frame(self.root)  # 创建文件区域
        file_frame.pack(fill=X, padx=10, pady=4)  # 布局文件区域
        Label(file_frame, text="批量数据导入：").pack(side=LEFT)  # 创建批量数据导入标签
        Entry(file_frame, textvariable=self.excel_path_var).pack(side=LEFT, fill=X, expand=True, padx=(4, 8))  # 创建Excel输入框
        Button(file_frame, text="选择文件", command=self.choose_excel).pack(side=RIGHT)  # 创建选择文件按钮
        field_frame = Frame(self.root)  # 创建字段区域
        field_frame.pack(fill=X, padx=10, pady=6)  # 布局字段区域
        field_head = Frame(field_frame)  # 创建字段头区域
        field_head.pack(fill=X)  # 布局字段头区域
        Label(field_head, text="字段配置（要求列只读，选填为空不发送，条件必填仅提示）：").pack(side=LEFT)  # 创建字段提示
        Button(field_head, text="导出Excel模板", command=self.export_template, width=16).pack(side=RIGHT, padx=(8, 0))  # 创建导出模板按钮
        Button(field_head, text="添加字段", command=self.add_field, width=12).pack(side=RIGHT)  # 创建添加字段按钮
        field_canvas_frame = Frame(field_frame, height=260)  # 创建字段滚动外框
        field_canvas_frame.pack(fill=X, pady=4)  # 布局字段滚动外框
        field_canvas_frame.pack_propagate(False)  # 固定字段区域高度
        self.field_canvas = Canvas(field_canvas_frame, highlightthickness=0)  # 创建字段画布
        self.field_scrollbar = Scrollbar(field_canvas_frame, orient="vertical", command=self.field_canvas.yview)  # 创建字段滚动条
        self.field_rows_frame = Frame(self.field_canvas)  # 创建字段行容器
        self.field_rows_frame.bind("<Configure>", lambda _event: self.field_canvas.configure(scrollregion=self.field_canvas.bbox("all")))  # 更新滚动区域
        self.field_canvas.create_window((0, 0), window=self.field_rows_frame, anchor="nw")  # 放入字段行容器
        self.field_canvas.configure(yscrollcommand=self.field_scrollbar.set)  # 绑定滚动条
        self.field_canvas.pack(side=LEFT, fill=BOTH, expand=True)  # 布局画布
        self.field_scrollbar.pack(side=RIGHT, fill="y")  # 布局滚动条
        self.field_canvas.bind("<Enter>", self.bind_field_mousewheel)  # 启用鼠标滚轮
        self.field_canvas.bind("<Leave>", self.unbind_field_mousewheel)  # 取消鼠标滚轮
        self.field_rows_frame.bind("<Enter>", self.bind_field_mousewheel)  # 启用鼠标滚轮
        self.field_rows_frame.bind("<Leave>", self.unbind_field_mousewheel)  # 取消鼠标滚轮
        action_frame = Frame(self.root)  # 创建操作区域
        action_frame.pack(fill=X, padx=10, pady=6)  # 布局操作区域
        self.start_button = Button(action_frame, text="开始执行", command=self.start_process, width=14)  # 创建开始按钮
        self.start_button.config(bg="#2f80ed", fg="#ffffff", activebackground="#1f6fd1", activeforeground="#ffffff")  # 设置主按钮样式
        self.start_button.pack(side=LEFT)  # 布局开始按钮
        self.pause_button = Button(action_frame, text="暂停执行", command=self.pause_process, width=14, state="disabled")  # 创建暂停按钮
        self.pause_button.pack(side=LEFT, padx=8)  # 布局暂停按钮
        self.stop_button = Button(action_frame, text="停止执行", command=self.stop_process, width=14, state="disabled")  # 创建停止按钮
        self.stop_button.pack(side=LEFT)  # 布局停止按钮
        Label(action_frame, textvariable=self.status_var).pack(side=LEFT, padx=8)  # 创建状态标签
        Button(action_frame, text="导出日志", command=self.export_log, width=12).pack(side=RIGHT)  # 创建导出日志按钮
        Button(action_frame, text="清空日志", command=self.clear_log, width=12).pack(side=RIGHT, padx=8)  # 创建清空日志按钮
        log_frame = Frame(self.root, height=160)  # 创建日志区域
        log_frame.pack(fill=X, padx=10, pady=4)  # 布局日志区域
        log_frame.pack_propagate(False)  # 固定日志高度
        Label(log_frame, text="调用详细日志：").pack(anchor="w")  # 创建日志标题
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap="word", height=6)  # 创建日志文本框
        self.log_text.pack(fill=BOTH, expand=True)  # 布局日志文本框
        failure_frame = Frame(self.root)  # 创建失败区域
        failure_frame.pack(fill=BOTH, expand=True, padx=10, pady=6)  # 布局失败区域
        failure_head = Frame(failure_frame)  # 创建失败头区域
        failure_head.pack(fill=X)  # 布局失败头区域
        Label(failure_head, text="失败行列表：").pack(side=LEFT)  # 创建失败标题
        Button(failure_head, text="导出失败行", command=self.export_failures, width=14).pack(side=RIGHT)  # 创建导出失败按钮
        Button(failure_head, text="清空失败行", command=self.clear_failures, width=14).pack(side=RIGHT, padx=8)  # 创建清空失败按钮
        failure_content = Frame(failure_frame)  # 创建失败内容区域
        failure_content.pack(fill=BOTH, expand=True, pady=4)  # 布局失败内容区域
        self.failure_tree = ttk.Treeview(failure_content, columns=("row", "summary", "reason"), show="headings", height=8)  # 创建失败表格
        self.failure_tree.heading("row", text="数据行")  # 设置数据行列
        self.failure_tree.heading("summary", text="字段值摘要")  # 设置摘要列
        self.failure_tree.heading("reason", text="失败原因")  # 设置原因列
        self.failure_tree.column("row", width=80, anchor="center")  # 设置数据行宽度
        self.failure_tree.column("summary", width=380, anchor="w")  # 设置摘要宽度
        self.failure_tree.column("reason", width=300, anchor="w")  # 设置原因宽度
        self.failure_tree.pack(side=LEFT, fill=BOTH, expand=True)  # 布局失败表格
        self.failure_tree.bind("<<TreeviewSelect>>", self.on_failure_selected)  # 绑定失败选择事件
        self.failure_detail_text = scrolledtext.ScrolledText(failure_content, wrap="word", width=55)  # 创建失败明细框
        self.failure_detail_text.pack(side=RIGHT, fill=BOTH, expand=True, padx=(8, 0))  # 布局失败明细框

    def show_network_settings(self) -> None:
        window = Toplevel(self.root)
        window.title("网络与 V4 签名设置")
        window.geometry("540x410")
        values = {}
        fields = [
            ("连接超时（秒）", "connect_timeout", False),
            ("读取超时（秒）", "read_timeout", False),
            ("App Key", "app_key", False),
            ("App Secret", "app_secret", True),
            ("请求来源", "source", False),
            ("API 版本", "api_version", False),
        ]
        for row_index, (label, key, secret) in enumerate(fields):
            Label(window, text=label).grid(row=row_index, column=0, sticky="e", padx=10, pady=7)
            variable = StringVar(value=str(NETWORK_SETTINGS[key]))
            values[key] = variable
            Entry(window, textvariable=variable, show="*" if secret else "").grid(
                row=row_index, column=1, sticky="ew", padx=10, pady=7
            )
        algorithm_var = StringVar(value=NETWORK_SETTINGS["algorithm"])
        Label(window, text="签名算法").grid(row=6, column=0, sticky="e", padx=10, pady=7)
        ttk.Combobox(
            window,
            textvariable=algorithm_var,
            values=["HMAC-SHA256", "HMAC-SHA512"],
            state="readonly",
        ).grid(row=6, column=1, sticky="ew", padx=10, pady=7)
        verify_var = BooleanVar(value=NETWORK_SETTINGS["verify_tls"])
        Checkbutton(window, text="校验 HTTPS 证书", variable=verify_var).grid(
            row=7, column=1, sticky="w", padx=10, pady=7
        )
        window.columnconfigure(1, weight=1)

        def save():
            try:
                NETWORK_SETTINGS["connect_timeout"] = max(0.1, float(values["connect_timeout"].get()))
                NETWORK_SETTINGS["read_timeout"] = max(0.1, float(values["read_timeout"].get()))
            except ValueError:
                messagebox.showerror("输入错误", "超时时间必须是数字", parent=window)
                return
            for key in ("app_key", "app_secret", "source", "api_version"):
                NETWORK_SETTINGS[key] = values[key].get().strip()
            NETWORK_SETTINGS["algorithm"] = algorithm_var.get()
            NETWORK_SETTINGS["verify_tls"] = verify_var.get()
            window.destroy()

        Button(window, text="保存", command=save, width=14).grid(row=8, column=1, sticky="e", padx=10, pady=15)

    def show_callback_server(self) -> None:
        window = Toplevel(self.root)
        window.title("本地 HTTP 回调模拟服务")
        window.geometry("1040x680")
        controls = Frame(window)
        controls.pack(fill=X, padx=10, pady=8)
        host_var = StringVar(value="0.0.0.0")
        port_var = StringVar(value="8090")
        mode_var = StringVar(value="成功")
        delay_var = StringVar(value="0")
        Label(controls, text="监听地址").pack(side=LEFT)
        Entry(controls, textvariable=host_var, width=13).pack(side=LEFT, padx=5)
        Label(controls, text="端口").pack(side=LEFT)
        Entry(controls, textvariable=port_var, width=7).pack(side=LEFT, padx=5)
        Label(controls, text="响应").pack(side=LEFT)
        ttk.Combobox(controls, textvariable=mode_var, values=["成功", "失败"], state="readonly", width=7).pack(side=LEFT, padx=5)
        Label(controls, text="延时(秒)").pack(side=LEFT)
        Entry(controls, textvariable=delay_var, width=7).pack(side=LEFT, padx=5)
        status_var = StringVar(value="未启动")
        Label(controls, textvariable=status_var).pack(side=RIGHT)
        tree = ttk.Treeview(window, columns=("time", "name", "task", "path"), show="headings", height=12)
        for key, title, width in (("time", "时间", 150), ("name", "回调", 180), ("task", "任务号", 180), ("path", "路径", 440)):
            tree.heading(key, text=title)
            tree.column(key, width=width, anchor="w")
        tree.pack(fill=X, padx=10, pady=4)
        detail = scrolledtext.ScrolledText(window, wrap="word")
        detail.pack(fill=BOTH, expand=True, padx=10, pady=6)
        record_map = {}

        def display_record(record):
            item = tree.insert("", END, values=(record["time"], record["name"], record["task_id"], record["path"]))
            record_map[item] = record

        def on_record(record):
            self.root.after(0, display_record, record)

        def start():
            try:
                delay = max(0, float(delay_var.get()))
                port = int(port_var.get())
                body = {"code": "0", "message": "成功"} if mode_var.get() == "成功" else {"code": "1", "message": "模拟失败"}
                status = 200 if mode_var.get() == "成功" else 500
                if self.callback_server:
                    self.callback_server.stop()
                self.callback_server = CallbackServer(host_var.get().strip(), port, on_record)
                self.callback_server.profile = ResponseProfile(status, delay, body)
                self.callback_server.start()
                status_var.set(f"运行中：http://127.0.0.1:{self.callback_server.port}")
            except (ValueError, OSError) as error:
                messagebox.showerror("启动失败", str(error), parent=window)

        def stop():
            if self.callback_server:
                self.callback_server.stop()
                self.callback_server = None
            status_var.set("已停止")

        def selected(_event=None):
            selection = tree.selection()
            if selection:
                detail.delete("1.0", END)
                detail.insert(END, json.dumps(record_map[selection[0]], ensure_ascii=False, indent=2))

        tree.bind("<<TreeviewSelect>>", selected)
        Button(controls, text="启动", command=start).pack(side=LEFT, padx=6)
        Button(controls, text="停止", command=stop).pack(side=LEFT)
        paths = {**V3_CALLBACK_PATHS, **V4_CALLBACK_PATHS}
        detail.insert(END, "支持路径（也接受任意自定义 POST 路径）：\n" + "\n".join(f"{path}  {name}" for path, name in paths.items()))
        window.protocol("WM_DELETE_WINDOW", window.destroy)

    def show_user_manual(self) -> None:  # 定义显示用户使用说明窗口的函数
        manual_window = Toplevel(self.root)  # 创建说明弹窗
        manual_window.title("用户使用说明")  # 设置说明窗口标题
        manual_window.geometry("900x700")  # 设置说明窗口大小
        manual_text = scrolledtext.ScrolledText(manual_window, wrap="word")  # 创建可滚动说明文本框
        manual_text.pack(fill=BOTH, expand=True, padx=10, pady=10)  # 布局说明文本框
        manual_text.insert(END, USER_MANUAL_TEXT)  # 写入内置说明文本
        manual_text.config(state="disabled")  # 设置说明文本为只读

    def get_current_presets(self) -> list[dict]:  # 定义获取当前版本预置函数
        return get_presets(self.version_var.get())  # 返回当前版本预置

    def update_api_url_from_parts(self, *_args) -> None:  # 定义刷新完整URL函数
        ip_text = self.ip_var.get().strip() or DEFAULT_IP  # 获取IP
        port_text = self.port_var.get().strip() or ("443" if self.version_var.get() == VERSION_4 else "8182")  # 获取端口
        suffix_text = self.suffix_var.get().strip().lstrip("/")  # 获取后缀
        if self.version_var.get() == VERSION_4 and suffix_text.startswith("wcs/"):
            self.api_url_var.set(f"https://{ip_text}:{port_text}/{suffix_text}")
        else:
            self.api_url_var.set(get_base_url(self.version_var.get(), ip_text, port_text) + suffix_text)  # 设置完整URL

    def apply_version_defaults(self) -> None:  # 定义应用版本默认值函数
        self.port_var.set("443" if self.version_var.get() == VERSION_4 else "8182")  # 设置默认端口
        presets = self.get_current_presets()  # 获取当前预置接口
        names = [preset["display_name"] for preset in presets]  # 获取展示名称
        self.api_combo.configure(values=names)  # 更新接口下拉框
        if names:  # 判断是否存在预置接口
            self.selected_api_var.set(names[0])  # 选中第一个接口
            self.load_preset_api(presets[0])  # 加载第一个接口

    def on_version_selected(self, _event=None) -> None:  # 定义版本切换事件函数
        if self.is_running:  # 判断是否正在运行
            if not messagebox.askyesno("确认", "切换版本将清空当前状态，是否继续？"):  # 弹出确认框
                return  # 用户取消时返回
            self.stop_event.set()  # 设置停止信号
            self.pause_event.set()  # 唤醒暂停
        self.clear_log()  # 清空日志
        self.clear_failures()  # 清空失败记录
        self.clear_field_rows()  # 清空字段配置
        self.apply_version_defaults()  # 应用版本默认配置

    def on_preset_api_selected(self, _event=None) -> None:  # 定义接口选择事件函数
        preset = next((item for item in self.get_current_presets() if item["display_name"] == self.selected_api_var.get()), None)  # 查找预置接口
        if preset:  # 判断是否找到
            self.load_preset_api(preset)  # 加载预置接口

    def load_preset_api(self, preset: dict) -> None:  # 定义加载预置接口函数
        self.suffix_var.set(preset["suffix"])  # 设置接口后缀
        self.clear_field_rows()  # 清空字段行
        sorted_fields = sorted(preset["fields"], key=lambda item: requirement_sort_key(item[1]))  # 按必填排序字段
        for index, (field_name, requirement) in enumerate(sorted_fields, start=1):  # 遍历字段
            nested_enabled, nested_path, child_name = split_field_for_ui(field_name)  # 拆分嵌套字段
            self.add_field(child_name, index, nested_enabled, nested_path, "", requirement)  # 添加字段行

    def add_field(self, default_field_name: str = "", default_column_index: int | None = None, default_nested_enabled: bool = False, default_nested_path: str = "", default_field_value: str = "", default_requirement: str = "选填") -> None:  # 定义添加字段行函数
        row_number = len(self.field_rows) + 1  # 获取行号
        row_frame = Frame(self.field_rows_frame, bd=1, relief="flat", bg="#ffffff")  # 创建行容器
        row_frame.pack(fill=X, pady=2, padx=(0, 4))  # 布局行容器
        column_var = StringVar(value=str(default_column_index or row_number))  # 创建列号变量
        nested_var = BooleanVar(value=default_nested_enabled)  # 创建嵌套变量
        nested_path_var = StringVar(value=default_nested_path)  # 创建嵌套层级变量
        field_var = StringVar(value=default_field_name)  # 创建字段名变量
        requirement_var = StringVar(value=default_requirement)  # 创建要求变量
        value_var = StringVar(value=default_field_value)  # 创建字段值变量
        drag_label = Label(row_frame, text="☰", cursor="fleur", fg="#4f6f9f", bg="#ffffff", font=("Segoe UI Symbol", 15), width=2)  # 创建拖动手柄
        drag_label.pack(side=LEFT, padx=(4, 6))  # 布局拖动手柄
        Label(row_frame, text="Excel列号").pack(side=LEFT, padx=(8, 2))  # 布局列号标签
        column_entry = Entry(row_frame, textvariable=column_var, width=7, justify="center")  # 创建列号输入框
        column_entry.pack(side=LEFT)  # 布局列号输入框
        nested_entry = Entry(row_frame, textvariable=nested_path_var, width=24, state="normal" if default_nested_enabled else "disabled")  # 创建嵌套层级输入框
        Checkbutton(row_frame, text="有嵌套", variable=nested_var, command=lambda entry=nested_entry, var=nested_var: self.toggle_nested_entry(entry, var)).pack(side=LEFT, padx=(8, 2))  # 布局嵌套复选框
        Label(row_frame, text="嵌套层级").pack(side=LEFT, padx=(6, 2))  # 布局嵌套标签
        nested_entry.pack(side=LEFT)  # 布局嵌套输入框
        Label(row_frame, text="字段名").pack(side=LEFT, padx=(8, 2))  # 布局字段名标签
        Entry(row_frame, textvariable=field_var, width=16).pack(side=LEFT)  # 布局字段名输入框
        Label(row_frame, text="要求").pack(side=LEFT, padx=(8, 2))  # 布局要求标签
        Entry(row_frame, textvariable=requirement_var, width=16, state="readonly").pack(side=LEFT)  # 布局要求只读框
        Label(row_frame, text="字段值").pack(side=LEFT, padx=(8, 2))  # 布局字段值标签
        value_entry = Entry(row_frame, textvariable=value_var, width=18, state="normal" if self.call_mode_var.get() == "单次" else "disabled")  # 创建字段值输入框
        value_entry.pack(side=LEFT)  # 布局字段值输入框
        Button(row_frame, text="删除", command=lambda frame=row_frame: self.remove_field(frame), width=7).pack(side=LEFT, padx=6)  # 创建删除按钮
        for child in row_frame.winfo_children():  # 遍历字段行控件
            if isinstance(child, (Label, Checkbutton)):  # 判断是否为文字类控件
                child.configure(bg="#ffffff")  # 统一字段行背景色
        row_data = {"frame": row_frame, "column_var": column_var, "column_entry": column_entry, "nested_var": nested_var, "nested_path_var": nested_path_var, "field_var": field_var, "requirement_var": requirement_var, "value_var": value_var, "value_entry": value_entry}  # 构造字段行数据
        self.field_rows.append(row_data)  # 保存字段行
        column_entry.bind("<FocusIn>", lambda _event, row=row_data: self.remember_column_number(row))  # 记录修改前列号
        column_entry.bind("<FocusOut>", lambda _event, row=row_data: self.apply_column_number(row))  # 失焦时应用列号排序
        column_entry.bind("<Return>", lambda _event, row=row_data: self.apply_column_number(row))  # 回车时应用列号排序
        drag_label.bind("<ButtonPress-1>", lambda event, frame=row_frame: self.start_drag_field_row(event, frame))  # 绑定拖动开始
        drag_label.bind("<B1-Motion>", self.drag_field_row)  # 绑定拖动移动
        drag_label.bind("<ButtonRelease-1>", self.end_drag_field_row)  # 绑定拖动结束

    def toggle_nested_entry(self, nested_entry: Entry, nested_var: BooleanVar) -> None:  # 定义嵌套输入框切换函数
        nested_entry.config(state="normal" if nested_var.get() else "disabled")  # 按勾选状态启用禁用

    def bind_field_mousewheel(self, _event) -> None:  # 定义绑定滚轮函数
        self.root.bind_all("<MouseWheel>", self.on_field_mousewheel)  # 绑定滚轮事件

    def unbind_field_mousewheel(self, _event) -> None:  # 定义解绑滚轮函数
        self.root.unbind_all("<MouseWheel>")  # 解绑滚轮事件

    def on_field_mousewheel(self, event) -> None:  # 定义滚轮处理函数
        self.field_canvas.yview_scroll(-1 if event.delta > 0 else 1, "units")  # 滚动字段画布

    def remember_column_number(self, row: dict) -> None:  # 定义记录列号函数
        row["previous_column_index"] = self.field_rows.index(row) + 1  # 保存当前字段位置

    def apply_column_number(self, row: dict) -> None:  # 定义应用列号自动排序函数
        if row not in self.field_rows:  # 判断字段行是否仍存在
            return  # 直接返回
        column_text = row["column_var"].get().strip()  # 获取用户输入列号
        current_index = self.field_rows.index(row)  # 获取当前字段索引
        if not column_text.isdigit():  # 判断是否为数字
            row["column_var"].set(str(current_index + 1))  # 恢复当前列号
            messagebox.showinfo("提示", "Excel列号请输入正整数。")  # 提示输入规则
            return  # 直接返回
        target_index = int(column_text) - 1  # 转为目标索引
        if target_index < 0 or target_index >= len(self.field_rows):  # 判断目标是否超出范围
            row["column_var"].set(str(current_index + 1))  # 恢复当前列号
            messagebox.showinfo("提示", f"Excel列号范围为 1～{len(self.field_rows)}。")  # 提示有效范围
            return  # 直接返回
        if target_index != current_index:  # 判断位置是否变化
            moving_row = self.field_rows.pop(current_index)  # 移出当前字段
            self.field_rows.insert(target_index, moving_row)  # 插入目标列号位置
            self.repack_field_rows()  # 重新布局并连续编号
        else:  # 处理位置未变化
            self.refresh_field_column_numbers()  # 恢复标准连续列号

    def refresh_field_column_numbers(self) -> None:  # 定义刷新列号函数
        for index, row in enumerate(self.field_rows, start=1):  # 遍历字段行
            row["column_var"].set(str(index))  # 设置列号

    def repack_field_rows(self) -> None:  # 定义重新布局字段行函数
        for row in self.field_rows:  # 遍历字段行
            row["frame"].pack_forget()  # 取消布局
        for row in self.field_rows:  # 遍历字段行
            row["frame"].pack(fill=X, pady=2, padx=(0, 4))  # 重新布局
        self.refresh_field_column_numbers()  # 刷新列号

    def start_drag_field_row(self, _event, row_frame: Frame) -> None:  # 定义开始拖动函数
        self.dragging_field_frame = row_frame  # 记录拖动行
        row_frame.config(relief="solid")  # 设置边框

    def drag_field_row(self, event) -> None:  # 定义拖动移动函数
        row_frame = getattr(self, "dragging_field_frame", None)  # 获取拖动行
        if row_frame is None:  # 判断是否无拖动行
            return  # 直接返回
        current_index = next((index for index, row in enumerate(self.field_rows) if row["frame"] == row_frame), None)  # 查找当前索引
        if current_index is None:  # 判断是否未找到
            return  # 直接返回
        target_index = current_index  # 初始化目标索引
        for index, row in enumerate(self.field_rows):  # 遍历字段行
            midpoint = row["frame"].winfo_rooty() + row["frame"].winfo_height() / 2  # 计算行中点
            if event.y_root > midpoint:  # 判断鼠标是否在中点下方
                target_index = index  # 更新目标索引
        if target_index != current_index:  # 判断索引是否变化
            moving_row = self.field_rows.pop(current_index)  # 移出当前行
            self.field_rows.insert(target_index, moving_row)  # 插入目标位置
            self.repack_field_rows()  # 重新布局

    def end_drag_field_row(self, _event) -> None:  # 定义结束拖动函数
        row_frame = getattr(self, "dragging_field_frame", None)  # 获取拖动行
        if row_frame is not None:  # 判断拖动行存在
            row_frame.config(relief="flat")  # 恢复边框
        self.dragging_field_frame = None  # 清空拖动行
        self.refresh_field_column_numbers()  # 刷新列号

    def clear_field_rows(self) -> None:  # 定义清空字段行函数
        for row in self.field_rows:  # 遍历字段行
            row["frame"].destroy()  # 销毁行
        self.field_rows.clear()  # 清空列表

    def remove_field(self, row_frame: Frame) -> None:  # 定义删除字段行函数
        if len(self.field_rows) <= 1:  # 判断是否最后一行
            messagebox.showinfo("提示", "至少需要保留一个字段。")  # 提示至少保留
            return  # 直接返回
        self.field_rows = [row for row in self.field_rows if row["frame"] != row_frame]  # 移除对应行
        row_frame.destroy()  # 销毁行
        self.repack_field_rows()  # 重新布局

    def get_field_configs(self) -> tuple[list[dict], str]:  # 定义获取字段配置函数
        configs = []  # 初始化字段配置
        seen_columns = set()  # 初始化列号集合
        seen_fields = set()  # 初始化字段名集合
        for row in self.field_rows:  # 遍历字段行
            column_text = row["column_var"].get().strip()  # 获取列号
            field_name = row["field_var"].get().strip()  # 获取字段名
            nested_enabled = row["nested_var"].get()  # 获取嵌套状态
            nested_path = row["nested_path_var"].get().strip()  # 获取嵌套层级
            requirement = row["requirement_var"].get().strip() or "选填"  # 获取字段要求
            if not column_text.isdigit():  # 判断列号是否数字
                return [], "Excel列号必须填写数字。"  # 返回错误
            if not field_name:  # 判断字段名是否为空
                return [], "字段名不能为空。"  # 返回错误
            if nested_enabled and not nested_path:  # 判断嵌套层级是否为空
                return [], "勾选有嵌套时，嵌套层级不能为空。"  # 返回错误
            column_index = int(column_text)  # 转为数字列号
            full_field_name = combine_nested_field_name(nested_path, field_name) if nested_enabled else field_name  # 组合完整字段名
            if full_field_name == "reqCode":  # 判断是否配置reqCode
                return [], "reqCode由程序自动生成，请不要手动配置。"  # 返回错误
            if column_index in seen_columns:  # 判断列号重复
                return [], f"Excel第{column_index}列被重复配置。"  # 返回错误
            if full_field_name in seen_fields:  # 判断字段重复
                return [], f"字段{full_field_name}被重复配置。"  # 返回错误
            seen_columns.add(column_index)  # 记录列号
            seen_fields.add(full_field_name)  # 记录字段名
            configs.append({"column_index": column_index, "field_name": full_field_name, "requirement": requirement})  # 添加配置
        return configs, ""  # 返回配置

    def get_single_call_row(self, field_configs: list[dict]) -> tuple[dict, str]:  # 定义获取单次调用数据函数
        values = {}  # 初始化字段值
        for row, config in zip(self.field_rows, field_configs):  # 遍历字段行
            field_name = config["field_name"]  # 获取字段名
            field_value = row["value_var"].get().strip()  # 获取字段值
            if is_strict_required(config.get("requirement", "选填")) and not field_value:  # 判断必填为空
                return {}, f"字段{field_name}为必填，当前为空。"  # 返回错误
            if field_value:  # 判断字段值非空
                values[field_name] = field_value  # 保存字段值
        return {"row_index": "单次", "values": values}, ""  # 返回单次数据

    def choose_excel(self) -> None:  # 定义选择Excel函数
        file_path = filedialog.askopenfilename(title="请选择Excel文件", filetypes=[("Excel文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")])  # 打开文件选择框
        if file_path:  # 判断是否选择文件
            self.excel_path_var.set(file_path)  # 设置文件路径

    def export_template(self) -> None:  # 定义导出模板函数
        field_configs, error = self.get_field_configs()  # 获取字段配置
        if error:  # 判断是否错误
            messagebox.showerror("错误", error)  # 弹出错误
            return  # 直接返回
        file_path = filedialog.asksaveasfilename(title="导出Excel模板", defaultextension=".xlsx", initialfile=f"RCS接口模板_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", filetypes=[("Excel文件", "*.xlsx")])  # 选择保存路径
        if not file_path:  # 判断是否取消
            return  # 直接返回
        workbook = Workbook()  # 创建工作簿
        worksheet = workbook.active  # 获取工作表
        worksheet.title = "模板"  # 设置工作表名
        max_column = max(config["column_index"] for config in field_configs)  # 获取最大列号
        headers = [""] * max_column  # 创建表头
        for config in field_configs:  # 遍历字段配置
            headers[config["column_index"] - 1] = config["field_name"]  # 写入表头
        worksheet.append(headers)  # 添加表头行
        workbook.save(file_path)  # 保存模板
        messagebox.showinfo("完成", f"Excel模板已导出：{file_path}")  # 提示完成

    def clear_log(self) -> None:  # 定义清空日志函数
        self.log_text.delete("1.0", END)  # 清空日志文本

    def export_log(self) -> None:  # 定义导出日志函数
        self.drain_log_queue()  # 先显示队列中尚未刷新的日志
        log_content = self.log_text.get("1.0", END).strip()  # 获取当前界面日志
        if not log_content:  # 判断是否无日志
            messagebox.showinfo("提示", "当前没有日志可以导出。")  # 提示无日志
            return  # 返回
        file_path = filedialog.asksaveasfilename(title="导出日志", defaultextension=".txt", initialfile=f"RCS接口日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt", filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")])  # 选择保存路径
        if not file_path:  # 判断是否取消
            return  # 返回
        try:  # 捕获文件保存异常
            Path(file_path).write_text(log_content + "\n", encoding="utf-8-sig")  # 保存界面日志
        except OSError as error:  # 捕获写入失败
            messagebox.showerror("错误", f"日志导出失败：{error}")  # 提示失败
            return  # 返回
        messagebox.showinfo("完成", f"日志已导出：{file_path}")  # 提示完成

    def add_log(self, text: str) -> None:  # 定义添加日志函数
        self.log_queue.put(text)  # 放入日志队列

    def drain_log_queue(self) -> None:  # 定义取出日志队列内容函数
        while not self.log_queue.empty():  # 循环处理日志
            text = self.log_queue.get()  # 获取日志
            self.log_text.insert(END, text + "\n")  # 插入日志
            self.log_text.see(END)  # 滚动到底部

    def flush_log_queue(self) -> None:  # 定义刷新日志函数
        self.drain_log_queue()  # 取出并显示待处理日志
        self.root.after(100, self.flush_log_queue)  # 下次刷新

    def clear_failures(self) -> None:  # 定义清空失败记录函数
        self.failed_records.clear()  # 清空失败记录
        for item_id in self.failure_tree.get_children():  # 遍历表格项
            self.failure_tree.delete(item_id)  # 删除表格项
        self.failure_detail_text.delete("1.0", END)  # 清空失败明细

    def build_failure_record(self, row: dict, message: str, request_text: str, response_text: str) -> dict:  # 定义构造失败记录函数
        summary = "，".join([f"{field}={value}" for field, value in row["values"].items()])  # 生成摘要
        return {"row_index": row["row_index"], "values": row["values"], "summary": summary, "message": message, "request_text": request_text, "response_text": response_text}  # 返回失败记录

    def add_failure_record(self, record: dict) -> None:  # 定义添加失败记录函数
        item_id = f"fail_{uuid.uuid4().hex}"  # 生成唯一项ID
        self.failed_records[item_id] = record  # 保存失败记录
        summary_display = record["summary"][:300] + "..." if len(record["summary"]) > 300 else record["summary"]  # 截断摘要
        self.failure_tree.insert("", END, iid=item_id, values=(record["row_index"], summary_display, record["message"]))  # 插入失败表格
        self.failure_tree.see(item_id)  # 滚动到失败项

    def on_failure_selected(self, _event) -> None:  # 定义失败项选择函数
        selected_items = self.failure_tree.selection()  # 获取选中项
        if not selected_items:  # 判断是否未选中
            return  # 直接返回
        record = self.failed_records.get(selected_items[0])  # 获取失败记录
        if not record:  # 判断记录是否不存在
            return  # 直接返回
        detail = f"数据行：{record['row_index']}\n字段值摘要：{record['summary']}\n失败原因：{record['message']}\n\n发送报文：\n{record['request_text']}\n\n应答报文：\n{record['response_text']}"  # 生成明细
        self.failure_detail_text.delete("1.0", END)  # 清空明细
        self.failure_detail_text.insert(END, detail)  # 写入明细

    def export_failures(self) -> None:  # 定义导出失败函数
        if not self.failed_records:  # 判断是否无失败
            messagebox.showinfo("提示", "当前没有失败行可以导出。")  # 提示无失败
            return  # 返回
        file_path = filedialog.asksaveasfilename(title="导出失败行", defaultextension=".xlsx", initialfile=f"失败行_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", filetypes=[("Excel文件", "*.xlsx")])  # 选择保存路径
        if not file_path:  # 判断是否取消
            return  # 返回
        workbook = Workbook()  # 创建工作簿
        worksheet = workbook.active  # 获取工作表
        worksheet.title = "失败行"  # 设置工作表名
        worksheet.append(["数据行", "字段值JSON", "失败原因", "发送报文", "应答报文"])  # 写入表头
        for record in self.failed_records.values():  # 遍历失败记录
            worksheet.append([record["row_index"], json.dumps(record["values"], ensure_ascii=False), record["message"], record["request_text"], record["response_text"]])  # 写入失败行
        workbook.save(file_path)  # 保存文件
        messagebox.showinfo("完成", f"失败行已导出：{file_path}")  # 提示完成

    def update_call_mode_ui(self) -> None:  # 定义更新调用方式界面函数
        is_single = self.call_mode_var.get() == "单次"  # 判断是否单次
        for row in self.field_rows:  # 遍历字段行
            row["value_entry"].config(state="normal" if is_single else "disabled")  # 启用或禁用字段值
        self.status_var.set("单次调用模式：请填写字段值。" if is_single else "批量调用模式：请选择Excel文件。")  # 更新状态

    def set_running(self, running: bool) -> None:  # 定义设置运行状态函数
        self.is_running = running  # 设置运行状态
        self.start_button.config(text="开始执行", state="disabled" if running else "normal")  # 设置开始按钮
        self.pause_button.config(state="normal" if running else "disabled")  # 设置暂停按钮
        self.stop_button.config(state="normal" if running else "disabled")  # 设置停止按钮
        self.is_paused = False  # 清除暂停状态

    def pause_process(self) -> None:  # 定义暂停函数
        if not self.is_running or self.is_paused:  # 判断是否不能暂停
            return  # 返回
        self.is_paused = True  # 设置暂停状态
        self.pause_event.clear()  # 清除执行事件
        self.start_button.config(text="继续执行", state="normal")  # 设置继续按钮
        self.pause_button.config(state="disabled")  # 禁用暂停按钮

    def resume_process(self) -> None:  # 定义继续函数
        self.is_paused = False  # 清除暂停状态
        self.pause_event.set()  # 允许继续
        self.start_button.config(text="开始执行", state="disabled")  # 禁用开始按钮
        self.pause_button.config(state="normal")  # 启用暂停按钮

    def stop_process(self) -> None:  # 定义停止函数
        self.stop_event.set()  # 设置停止信号
        self.pause_event.set()  # 唤醒暂停
        self.status_var.set("正在停止，请等待当前请求结束。")  # 更新状态

    def wait_if_paused_or_stopped(self) -> bool:  # 定义暂停停止检查函数
        while not self.pause_event.is_set():  # 判断是否暂停
            if self.stop_event.is_set():  # 判断是否停止
                return False  # 返回不继续
            time.sleep(0.1)  # 等待
        return not self.stop_event.is_set()  # 返回是否继续

    def start_process(self) -> None:  # 定义开始执行函数
        if self.is_running and self.is_paused:  # 判断是否暂停中
            self.resume_process()  # 继续执行
            return  # 返回
        if self.is_running:  # 判断是否运行中
            return  # 返回
        api_url = self.api_url_var.get().strip()  # 获取完整URL
        field_configs, error = self.get_field_configs()  # 获取字段配置
        is_single = self.call_mode_var.get() == "单次"  # 判断是否单次
        if not api_url:  # 判断URL是否为空
            messagebox.showerror("错误", "完整URL不能为空。")  # 提示错误
            return  # 返回
        if error:  # 判断字段错误
            messagebox.showerror("错误", error)  # 提示错误
            return  # 返回
        if is_single:  # 判断单次模式
            single_row, single_error = self.get_single_call_row(field_configs)  # 获取单次数据
            if single_error:  # 判断单次错误
                messagebox.showerror("错误", single_error)  # 提示错误
                return  # 返回
        else:  # 处理批量模式
            excel_path = Path(self.excel_path_var.get().strip().strip('"'))  # 获取Excel路径
            if not excel_path.exists():  # 判断Excel是否存在
                messagebox.showerror("错误", f"Excel文件不存在：{excel_path}")  # 提示错误
                return  # 返回
            if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:  # 判断格式
                messagebox.showerror("错误", "仅支持.xlsx或.xlsm格式的Excel文件。")  # 提示错误
                return  # 返回
        self.clear_log()  # 清空日志
        self.clear_failures()  # 清空失败记录
        self.stop_event.clear()  # 清除停止信号
        self.pause_event.set()  # 允许执行
        self.set_running(True)  # 设置运行中
        if is_single:  # 判断单次模式
            self.worker_thread = threading.Thread(target=self.process_single_call, args=(api_url, field_configs, single_row), daemon=True)  # 创建单次线程
        else:  # 处理批量模式
            self.worker_thread = threading.Thread(target=self.process_excel, args=(excel_path, api_url, field_configs), daemon=True)  # 创建批量线程
        self.worker_thread.start()  # 启动线程

    def process_single_call(self, api_url: str, field_configs: list[dict], row: dict) -> None:  # 定义单次处理函数
        try:  # 捕获异常
            self.add_log(f"接口URL：{api_url}")  # 输出URL
            success, message, payload, response_text = post_api_request(self.version_var.get(), api_url, row)  # 调用接口
            request_text = json.dumps(payload, ensure_ascii=False, indent=2)  # 格式化请求
            response_display = self.format_response_text(response_text)  # 格式化响应
            self.add_log(f"发送报文：\n{request_text}")  # 输出请求
            self.add_log(f"应答报文：\n{response_display}")  # 输出响应
            if not success:  # 判断失败
                self.root.after(0, lambda: self.add_failure_record(self.build_failure_record(row, message, request_text, response_display)))  # 添加失败记录
                self.finish_process(False, "单次调用失败。")  # 结束失败
                return  # 返回
            self.finish_process(True, "单次调用成功。")  # 结束成功
        except Exception as error:  # 捕获异常
            self.add_log(f"程序异常：{error}")  # 输出异常
            self.finish_process(False, "程序异常，处理已停止。")  # 结束失败

    def process_excel(self, excel_path: Path, api_url: str, field_configs: list[dict]) -> None:  # 定义批量处理函数
        try:  # 捕获异常
            rows = read_excel_rows(excel_path, field_configs)  # 读取Excel
            if not rows:  # 判断无数据
                self.finish_process(False, "未读取到可处理数据。")  # 结束失败
                return  # 返回
            failed_count = 0  # 初始化失败数
            for position, row in enumerate(rows, start=1):  # 遍历数据行
                if not self.wait_if_paused_or_stopped():  # 判断是否停止
                    self.finish_process(False, "已停止执行。")  # 结束执行
                    return  # 返回
                validation_error = validate_row(row, field_configs)  # 校验数据行
                if validation_error:  # 判断校验失败
                    failed_count += 1  # 累加失败数
                    failure_record = self.build_failure_record(row, validation_error, "<未发送，数据校验未通过>", "<无应答报文>")  # 构造失败记录
                    self.root.after(0, lambda record=failure_record: self.add_failure_record(record))  # 添加失败记录
                    continue  # 继续下一行
                success, message, payload, response_text = post_api_request(self.version_var.get(), api_url, row)  # 调用接口
                request_text = json.dumps(payload, ensure_ascii=False, indent=2)  # 格式化请求
                response_display = self.format_response_text(response_text)  # 格式化响应
                self.add_log(f"第 {position}/{len(rows)} 条，Excel第 {row['row_index']} 行")  # 输出进度
                self.add_log(f"发送报文：\n{request_text}")  # 输出请求
                self.add_log(f"应答报文：\n{response_display}")  # 输出响应
                if self.stop_event.is_set():  # 判断停止信号
                    self.finish_process(False, "已停止执行。")  # 结束执行
                    return  # 返回
                if not success:  # 判断接口失败
                    failed_count += 1  # 累加失败数
                    failure_record = self.build_failure_record(row, message, request_text, response_display)  # 构造失败记录
                    self.root.after(0, lambda record=failure_record: self.add_failure_record(record))  # 添加失败记录
                time.sleep(0.1)  # 短暂等待
            self.finish_process(failed_count == 0, f"全部完成，失败 {failed_count} 条。")  # 结束处理
        except Exception as error:  # 捕获异常
            self.add_log(f"程序异常：{error}")  # 输出异常
            self.finish_process(False, "程序异常，处理已停止。")  # 结束失败

    def format_response_text(self, response_text: str) -> str:  # 定义响应格式化函数
        if not response_text:  # 判断响应为空
            return "<无响应内容>"  # 返回占位文本
        try:  # 尝试解析JSON
            return json.dumps(json.loads(response_text), ensure_ascii=False, indent=2)  # 返回格式化JSON
        except json.JSONDecodeError:  # 捕获解析失败
            return response_text  # 返回原文

    def finish_process(self, success: bool, message: str) -> None:  # 定义结束处理函数
        self.root.after(0, lambda: self.status_var.set(message))  # 更新状态
        self.root.after(0, lambda: self.set_running(False))  # 恢复按钮
        if success:  # 判断成功
            self.root.after(0, lambda: messagebox.showinfo("完成", message))  # 弹出成功
        else:  # 处理失败
            self.root.after(0, lambda: messagebox.showwarning("提醒", message))  # 弹出提醒

    def get_raw_field_configs_for_save(self) -> list[dict]:  # 定义获取保存字段函数
        return [{"column_index": row["column_var"].get().strip(), "field_name": row["field_var"].get().strip(), "nested_enabled": row["nested_var"].get(), "nested_path": row["nested_path_var"].get().strip(), "requirement": row["requirement_var"].get().strip(), "field_value": row["value_var"].get()} for row in self.field_rows]  # 返回字段配置

    def save_config(self) -> None:  # 定义保存配置函数
        config = {"version": self.version_var.get(), "ip": self.ip_var.get(), "port": self.port_var.get(), "suffix": self.suffix_var.get(), "api_url": self.api_url_var.get(), "selected_api": self.selected_api_var.get(), "call_mode": self.call_mode_var.get(), "excel_path": self.excel_path_var.get(), "fields": self.get_raw_field_configs_for_save()}  # 组装配置
        try:  # 捕获写入异常
            get_config_path().write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")  # 写入配置
        except OSError as error:  # 捕获IO异常
            self.add_log(f"保存配置失败：{error}")  # 输出保存失败

    def load_config_or_default(self) -> None:  # 定义加载配置或默认值函数
        config_path = get_config_path()  # 获取新版配置路径
        if not config_path.exists() and get_legacy_config_path().exists():  # 判断是否只有旧版配置
            config_path = get_legacy_config_path()  # 兼容读取旧版配置
        if config_path.exists():  # 判断配置是否存在
            try:  # 捕获读取异常
                config = json.loads(config_path.read_text(encoding="utf-8"))  # 读取配置
                self.version_var.set(config.get("version", DEFAULT_VERSION))  # 恢复版本
                self.ip_var.set(config.get("ip", DEFAULT_IP))  # 恢复IP
                self.port_var.set(config.get("port", "443" if self.version_var.get() == VERSION_4 else "8182"))  # 恢复端口
                self.suffix_var.set(config.get("suffix", ""))  # 恢复后缀
                self.api_url_var.set(config.get("api_url", ""))  # 恢复完整URL
                saved_call_mode = config.get("call_mode", "单次")  # 读取保存的调用方式
                if saved_call_mode == "single":  # 兼容旧配置中的单次英文值
                    saved_call_mode = "单次"  # 转换为中文单次
                if saved_call_mode == "batch":  # 兼容旧配置中的批量英文值
                    saved_call_mode = "批量"  # 转换为中文批量
                self.call_mode_var.set(saved_call_mode if saved_call_mode in {"单次", "批量"} else "单次")  # 恢复调用方式
                self.excel_path_var.set(config.get("excel_path", ""))  # 恢复Excel路径
                self.refresh_api_options()  # 刷新接口选项
                self.selected_api_var.set(config.get("selected_api", ""))  # 恢复接口选择
                self.clear_field_rows()  # 清空字段行
                for field in config.get("fields", []):  # 遍历保存字段
                    column_text = str(field.get("column_index", len(self.field_rows) + 1)).strip()  # 获取列号文本
                    column_index = int(column_text) if column_text.isdigit() else len(self.field_rows) + 1  # 转为列号
                    self.add_field(field.get("field_name", ""), column_index, bool(field.get("nested_enabled", False)), field.get("nested_path", ""), field.get("field_value", ""), field.get("requirement", "选填"))  # 恢复字段行
                if not self.field_rows:  # 判断没有字段
                    self.apply_version_defaults()  # 应用默认值
                self.update_call_mode_ui()  # 刷新调用方式
                return  # 返回
            except (OSError, json.JSONDecodeError):  # 捕获配置异常
                pass  # 忽略异常
        self.apply_version_defaults()  # 应用默认版本

    def refresh_api_options(self) -> None:  # 定义刷新接口选项函数
        names = [preset["display_name"] for preset in self.get_current_presets()]  # 获取接口名称
        self.api_combo.configure(values=names)  # 更新下拉选项

    def on_close(self) -> None:  # 定义关闭窗口函数
        self.save_config()  # 保存配置
        if self.callback_server:
            self.callback_server.stop()
        if self.is_running:  # 判断是否运行中
            self.stop_event.set()  # 设置停止
            self.pause_event.set()  # 唤醒暂停
        self.root.destroy()  # 销毁窗口


def main() -> None:  # 定义主入口函数
    root = Tk()  # 创建主窗口
    RcsApiCallTool(root)  # 创建工具界面
    root.mainloop()  # 启动事件循环


if __name__ == "__main__":  # 判断是否直接运行
    main()  # 调用主函数
